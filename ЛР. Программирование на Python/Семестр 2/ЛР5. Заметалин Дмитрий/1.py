import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

wb = openpyxl.load_workbook('10.xlsx')
sheet = wb.active

s = Side(border_style="double")

n = sheet.max_row
m = sheet.max_column
count = 0
c = 0
yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
green_fill = PatternFill(start_color='008000', end_color='008000', fill_type='solid')


for row in sheet.iter_rows(min_row=1,
                           max_row=n, 
                           min_col=1,
                           max_col=m,
                           values_only=True):
    c += 1
    maxi = max(row)
    idx = "E"+str(c)
    if (sum(row) - maxi) < maxi and row[0] + row[1] != row[2] + row[3] and row[0] + row[2] != row[1] + row[3] and row[0] + row[3] != row[1] + row[2]:
        count += 1
        for row1 in sheet[f'A{c}:D{c}']:
            for cell in row1:
                cell.fill = green_fill

    else:
        for row1 in sheet[f'A{c}:D{c}']:
            for cell in row1:
                cell.fill = yellow_fill

    sheet[idx] = count

    if c == 1:
        sheet[f"A{c}"].border = Border(left=s, top=s)
        sheet[f"D{c}"].border = Border(right=s, top=s)
        for row in sheet[f"B{c}:C{c}"]:
            for cell in row:
                cell.border = Border(top=s)

    elif c == n:
        sheet[f"A{c}"].border = Border(left=s, bottom=s)
        sheet[f"D{c}"].border = Border(right=s, bottom=s)
        for row in sheet[f"B{c}:C{c}"]:
            for cell in row:
                cell.border = Border(bottom=s)

    else:
        sheet[f"A{c}"].border = Border(left=s)
        sheet[f"D{c}"].border = Border(right=s)


print(count)
wb.save('10.xlsx')